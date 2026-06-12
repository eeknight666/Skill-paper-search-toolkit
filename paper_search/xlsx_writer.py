"""Excel 导出与 PDF 批量下载"""

import re
import asyncio
import logging
from pathlib import Path

import httpx
from .base import Paper

logger = logging.getLogger(__name__)

EXCEL_HEADERS = [
    "论文名", "发表时间", "期刊/会议", "CCF等级", "作者",
    "论文分类", "简要描述", "相关性", "论文下载链接",
]


def export_xlsx(papers: list[Paper], filepath: str | Path) -> None:
    """将论文列表导出为 Excel 筛选表。

    自动按相关性评分降序排列，带颜色标记、冻结首行、自动筛选。
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("请先安装 openpyxl: pip install openpyxl")
        return

    filepath = Path(filepath)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "论文筛选结果"

    # 表头样式
    header_fill = openpyxl.styles.PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = openpyxl.styles.Font(color="FFFFFF", bold=True, size=11)

    for col, h in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    # 相关性颜色
    color_map = {
        "强相关": ("C6EFCE", "006100"),
        "相关": ("BDD7EE", "1F4E79"),
        "弱相关": ("FFF2CC", "9C6500"),
        "一般": ("F2F2F2", "595959"),
        "无关": ("F4B4C2", "9C0006"),
    }

    # 按评分降序
    papers_sorted = sorted(papers, key=lambda p: p.relevance_score, reverse=True)

    for row, p in enumerate(papers_sorted, 2):
        for col, val in enumerate(p.to_row(), 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=True)

        # 相关性列着色
        rel_cell = ws.cell(row=row, column=8)
        if p.relevance in color_map:
            bg, _ = color_map[p.relevance]
            rel_cell.fill = openpyxl.styles.PatternFill(
                start_color=bg, end_color=bg, fill_type="solid"
            )

    # 列宽
    col_widths = [45, 12, 28, 10, 35, 22, 60, 10, 50]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(papers_sorted) + 1}"

    wb.save(filepath)
    logger.info(f"Excel 已导出: {filepath} ({len(papers_sorted)} 条)")


async def download_pdfs(
    papers: list[Paper],
    output_dir: str | Path,
    concurrency: int = 3,
    min_size_bytes: int = 10000,
) -> tuple[int, int]:
    """批量下载论文 PDF。

    Args:
        papers: 论文列表（需要有 pdf_url）
        output_dir: 输出目录
        concurrency: 并发下载数
        min_size_bytes: 最小文件大小（过滤无效下载）

    Returns:
        (成功数, 总数)
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    sem = asyncio.Semaphore(concurrency)

    total = len([p for p in papers if p.pdf_url])
    success = 0

    async def _download_one(paper: Paper) -> bool:
        if not paper.pdf_url:
            return False
        # 安全文件名
        safe = re.sub(r'[\\/*?:"<>|]', "", paper.title)[:80].strip()
        fname = f"{paper.arxiv_id or safe}.pdf"
        fpath = output_dir / fname
        if fpath.exists():
            logger.info(f"  已存在: {fname}")
            return True
        async with sem:
            try:
                async with httpx.AsyncClient(
                    timeout=120.0, follow_redirects=True
                ) as client:
                    resp = await client.get(paper.pdf_url)
                    if resp.status_code == 200 and len(resp.content) > min_size_bytes:
                        fpath.write_bytes(resp.content)
                        logger.info(f"  OK: {fname} ({len(resp.content) // 1024} KB)")
                        return True
                    else:
                        logger.warning(f"  FAIL({resp.status_code}): {paper.title[:60]}")
                        return False
            except Exception as e:
                logger.warning(f"  ERR: {paper.title[:60]} - {e}")
                return False

    tasks = [_download_one(p) for p in papers if p.pdf_url]
    if tasks:
        results = await asyncio.gather(*tasks)
        success = sum(results)

    logger.info(f"PDF 下载: {success}/{total}")
    return success, total
